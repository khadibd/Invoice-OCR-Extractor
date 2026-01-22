import pandas as pd
import os
from openpyxl.styles import Border, Side
from openpyxl import load_workbook


class Exporter:

    @staticmethod
    def to_excel(all_fields, all_items, output_file):
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        # Convert to DataFrames
        df_fields = pd.DataFrame(all_fields)
        df_items = pd.DataFrame(all_items)

        # Write Excel first (pandas)
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df_fields.to_excel(writer, sheet_name="Invoices", index=False)
            df_items.to_excel(writer, sheet_name="Line_Items", index=False)

        # Apply borders (openpyxl)
        Exporter._apply_borders(output_file, "Invoices")
        Exporter._apply_borders(output_file, "Line_Items")

    @staticmethod
    def _apply_borders(excel_path, sheet_name):
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column
        ):
            for cell in row:
                cell.border = thin_border

        wb.save(excel_path)
